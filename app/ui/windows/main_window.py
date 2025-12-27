import os
import json
import time
from typing import List, Dict
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QStackedWidget,
    QSplitter, QScrollArea, QFrame, QApplication, QMessageBox,
    QLabel, QDialog
)
from PyQt6.QtCore import Qt, QTimer
from PyQt6.QtGui import QCursor

from app.core.controller import ParserController
from app.core.memory import MemoryManager
from app.core.telegram_notifier import TelegramNotifier
from app.core.tracker import AdTracker
from app.config import BASE_APP_DIR, RESULTS_DIR
from app.ui.widgets.ai_memory_panel import AIMemoryPanel
from app.core.ai.chunk_cultivation import ChunkCultivationManager
from app.ui.pages.analytics import AnalyticsWidget
from app.ui.windows.search_widget import SearchWidget
from app.ui.windows.controls_widget import ControlsWidget
from app.ui.windows.queue_state_manager import QueueStateManager
from app.ui.windows.settings_manager import SettingsDialog
from app.ui.widgets.results_area import ResultsAreaWidget
from app.ui.widgets.progress_and_logs_panel import ProgressAndLogsPanel
from app.ui.styles import Components, Palette, Spacing, Typography
from app.ui.widgets.category_selection_dialog import CategorySelectionDialog
from app.core.log_manager import logger

class ScanPromptDialog(QDialog):
    def __init__(self, queue_name: str, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Проверка категорий")
        self.setFixedWidth(500)
        self.setStyleSheet(Components.dialog())
        
        layout = QVBoxLayout(self)
        layout.setSpacing(Spacing.MD)
        layout.setContentsMargins(Spacing.LG, Spacing.LG, Spacing.LG, Spacing.LG)
        
        # Иконка или заголовок
        title = QLabel(f"Очередь: {queue_name}")
        title.setStyleSheet(Components.section_title())
        layout.addWidget(title)
        
        # Сообщение
        msg = QLabel(
            f"Обнаружено отсутствие отсканированных категорий поиска для очереди <b>{queue_name}</b>.<br><br>"
            "Это можно сделать сейчас, чтобы улучшить результаты поиска нужных объявлений, "
            "после чего парсер сразу начнет работу.<br><br>"
            "Либо процесс поиска будет запущен по одной «авто-категории», который выберет "
            "главную найденную категорию для поискового запроса."
        )
        msg.setWordWrap(True)
        msg.setStyleSheet(f"color: {Palette.TEXT}; font-size: 14px; line-height: 1.4;")
        layout.addWidget(msg)
        
        # Кнопки
        btn_layout = QVBoxLayout()
        btn_layout.setSpacing(Spacing.SM)
        
        self.btn_scan = QPushButton("Отсканировать и выбрать категории")
        self.btn_scan.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_scan.setStyleSheet(Components.start_button()) # Зеленая/Яркая кнопка
        self.btn_scan.setFixedHeight(45)
        
        self.btn_auto = QPushButton("Продолжить с авто-категорией")
        self.btn_auto.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_auto.setStyleSheet(Components.nav_button()) # Нейтральная кнопка
        self.btn_auto.setFixedHeight(40)
        
        self.btn_cancel = QPushButton("Отмена")
        self.btn_cancel.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_cancel.setStyleSheet(Components.stop_button()) # Красная/Серая кнопка
        
        btn_layout.addWidget(self.btn_scan)
        btn_layout.addWidget(self.btn_auto)
        btn_layout.addWidget(self.btn_cancel)
        
        layout.addLayout(btn_layout)
        
        self.btn_scan.clicked.connect(lambda: self.done(10))
        self.btn_auto.clicked.connect(lambda: self.done(20))
        self.btn_cancel.clicked.connect(self.reject)

class MainWindow(QWidget):
    def __init__(self):
        super().__init__()

        self.memory_manager = MemoryManager()
        self.controller = ParserController(memory_manager=self.memory_manager)
        self.controller.set_progress_callback(self._on_parser_progress)
        self.controller.ai_result_ready.connect(self._on_ai_result_with_memory)

        self._neuro_filtered: Dict[int, List[Dict]] = {}
        
        self.queue_manager = QueueStateManager()
        self.current_results = []
        self.current_json_file = None
        self.is_sequence_running = False
        self._is_programmatic_update = False
        self.cnt_parser = 0
        self.cnt_neuro = 0
        self.parser_progress_timer = None
        self.current_search_mode = "full"
        self.app_settings = self._load_settings()
        tg_token = self.app_settings.get("telegram_token", "")
        tg_chat_id = self.app_settings.get("telegram_chat_id", "")
        self.notifier = TelegramNotifier(tg_token, tg_chat_id)
        self.tracker = AdTracker(self.app_settings, self.notifier)
        self.tracker.item_updated.connect(self._on_tracker_item_updated)
        self.tracker.start()
        self.init_ui()
        self._connect_signals()
        self._restore_queues_from_state()
        self._start_timers()
        self.rag_rebuild_timer = QTimer(self)
        self.rag_rebuild_timer.timeout.connect(self.rebuild_rag_cache)
        self.rag_rebuild_timer.start(600000)
        self._validation_queue = []
        self._validating_index = None
        QTimer.singleShot(100, self._check_ai_availability)
        QTimer.singleShot(0, self.apply_initial_geometry)
        self._load_queue_to_ui(0)

        self.controller.ensure_ai_manager()
    
        self.controller.chunk_manager = ChunkCultivationManager(
            memory_manager=self.memory_manager,
            ai_manager=self.controller.ai_manager
        )

        if hasattr(self, 'memory_panel'):
            self.memory_panel.set_managers(
                self.memory_manager,
                self.controller.chunk_manager
            )

    def init_ui(self):
        self.setWindowTitle("Avito Assist")
        self.setStyleSheet(Components.main_window())
        self.setObjectName("ParserWidget")
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        self._init_top_bar(main_layout)
        self._init_navigation(main_layout)
        self.stack = QStackedWidget()
        main_layout.addWidget(self.stack)
        self.parser_page = self._create_parser_page()
        self.stack.addWidget(self.parser_page)
        self.analytics_page = self.create_analytics_page()
        self.stack.addWidget(self.analytics_page)
        self.memory_panel = AIMemoryPanel()
        self.stack.addWidget(self.memory_panel)
        self.memory_panel.load_instructions_from_disk()

        if not self.controller.chunk_manager:
            self.controller.chunk_manager = ChunkCultivationManager(
                memory_manager=self.memory_manager,
                ai_manager=self.controller.ai_manager
            )
            
        # Связываем панель с менеджерами
        self.memory_panel.set_managers(
            self.memory_manager,
            self.controller.chunk_manager
        )
        
        # Подключаем кнопку "Актуализировать" из панели к контроллеру
        self.memory_panel.update_memory_requested.connect(
            self.controller.start_cultivation
        )

        self.btn_parser.setChecked(True)
        self.stack.setCurrentIndex(0)

    def _init_top_bar(self, parent_layout):
        top_bar = QWidget()
        top_bar.setStyleSheet(f"background-color: {Palette.BG_DARK}; border-bottom: 1px solid {Palette.BORDER_SOFT};")
        top_layout = QHBoxLayout(top_bar)
        top_layout.setContentsMargins(Spacing.LG, Spacing.SM, Spacing.LG, Spacing.SM)
        
        title = QLabel("AVITO ASSIST")
        title.setStyleSheet(f"color: {Palette.PRIMARY}; font-weight: bold; font-size: 16px; border: none;")
        top_layout.addWidget(title)
        top_layout.addStretch()
        
        self.settings_button = QPushButton("⚙ Настройки")
        self.settings_button.setCursor(Qt.CursorShape.PointingHandCursor)
        self.settings_button.setStyleSheet(f"""
            QPushButton {{
                background: transparent;
                color: {Palette.TEXT_MUTED};
                border: 1px solid {Palette.BORDER_SOFT};
                border-radius: {Spacing.RADIUS_NORMAL}px;
                padding: {Spacing.XS}px {Spacing.MD}px;
            }}
            QPushButton:hover {{
                color: {Palette.TEXT};
                border-color: {Palette.TEXT_MUTED};
                background: {Palette.BG_DARK_2};
            }}
        """)
        self.settings_button.clicked.connect(self._on_settings_clicked)
        top_layout.addWidget(self.settings_button)
        parent_layout.addWidget(top_bar)

    def _init_navigation(self, parent_layout):
        nav_container = QWidget()
        nav_container.setStyleSheet(f"background-color: {Palette.BG_DARK};")
        nav_layout = QHBoxLayout(nav_container)
        nav_layout.setContentsMargins(Spacing.LG, Spacing.SM, Spacing.LG, Spacing.SM)
        nav_layout.setSpacing(Spacing.MD)
        
        # 1. Создаем кнопки (Порядок: Парсер -> Память -> Аналитика)
        self.btn_parser = QPushButton("ПАРСЕР")
        self.btn_memory = QPushButton("НЕЙРОСЕТЬ")
        self.btn_analytics = QPushButton("АНАЛИТИКА")
        
        # 2. Применяем единый стиль ко всем
        for btn in (self.btn_parser, self.btn_memory, self.btn_analytics):
            btn.setCheckable(True)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.setFixedHeight(36)
            btn.setStyleSheet(Components.nav_button()) # Используем тот же стиль, что и у Парсера
            
        # 3. Добавляем в лайаут в нужном порядке
        nav_layout.addWidget(self.btn_parser)
        nav_layout.addWidget(self.btn_memory)
        nav_layout.addWidget(self.btn_analytics)
        nav_layout.addStretch()
        
        parent_layout.addWidget(nav_container)
        
        # 4. Настраиваем переключение (индексы поменялись)
        self.btn_parser.clicked.connect(lambda: self._switch_page(0))
        self.btn_memory.clicked.connect(lambda: self._switch_page(2))    # Память теперь 2-й визуально, но индекс стека зависит от init_ui
        self.btn_analytics.clicked.connect(lambda: self._switch_page(1))

    def _create_parser_page(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        top_scroll = QScrollArea()
        top_scroll.setWidgetResizable(True)
        top_scroll.setFrameShape(QFrame.Shape.NoFrame)
        top_scroll.setStyleSheet(Components.scroll_area())
        top_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        top_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)

        top_content = QWidget()
        top_layout = QVBoxLayout(top_content)
        top_layout.setContentsMargins(Spacing.MD, Spacing.MD, Spacing.MD, Spacing.MD)
        top_layout.setSpacing(Spacing.GAP_SECTION)

        self.search_widget = SearchWidget()
        top_layout.addWidget(self.search_widget)

        self.controls_widget = ControlsWidget()

        split_saved = self.app_settings.get("global_split_results", False)
        if hasattr(self.controls_widget, "split_results_sw"):
            self.controls_widget.split_results_sw.setChecked(split_saved)

        top_layout.addWidget(self.controls_widget)

        # --- STATS CONTAINER (MODIFIED) ---
        stats_container = QWidget()
        stats_layout = QHBoxLayout(stats_container)
        stats_layout.setContentsMargins(0, 0, 0, 0)
        stats_layout.setSpacing(Spacing.MD)
        
        # NEW WIP PLACEHOLDER
        wip_frame = QFrame()
        wip_frame.setStyleSheet(f"""
            QFrame {{
                background-color: {Palette.BG_DARK_2};
                border: 1px dashed {Palette.BORDER_SOFT};
                border-radius: {Spacing.RADIUS_NORMAL}px;
            }}
        """)
        wip_frame.setFixedHeight(120) # Приличный размер
        wip_layout = QVBoxLayout(wip_frame)
        wip_label = QLabel("WIP (Work In Progress)")
        wip_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        wip_label.setStyleSheet(Typography.style(
            family=Typography.MONO, size=Typography.SIZE_XL, 
            weight=Typography.WEIGHT_BOLD, color=Palette.TEXT_MUTED
        ))
        wip_layout.addWidget(wip_label)
        
        stats_layout.addWidget(wip_frame)

        self.search_widget.attach_ai_stats(stats_container)

        self.progress_panel = ProgressAndLogsPanel()
        self.controls_widget.attach_progress_panel(self.progress_panel)

        top_scroll.setWidget(top_content)
        self.results_area = ResultsAreaWidget(self)

        self._refresh_merge_targets()

        splitter = QSplitter(Qt.Orientation.Vertical)
        splitter.setHandleWidth(9)
        splitter.addWidget(top_scroll)
        splitter.addWidget(self.results_area)
        splitter.setStretchFactor(0, 3)
        splitter.setStretchFactor(1, 7)
        splitter.setCollapsible(0, True)
        splitter.setSizes([600, 300])

        layout.addWidget(splitter)
        return page

    def create_analytics_page(self) -> QWidget:
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(Spacing.LG, Spacing.LG, Spacing.LG, Spacing.LG)
        
        # ИЗМЕНИ эту строку - передай controller
        self.analytics_widget = AnalyticsWidget(self.memory_manager, self.controller)
        layout.addWidget(self.analytics_widget)

        #self.database_tab = DatabaseTab(self.memory_manager)
        #self.stack.addWidget(self.database_tab)
        
        return page

    def _connect_signals(self):
        self.controller.parser_started.connect(self._on_parser_started_logic)
        self.controller.progress_updated.connect(self._on_parser_progress)
        self.controller.queue_finished.connect(self._on_queue_finished)
        self.controller.parser_finished.connect(self._on_parsing_finished)
        self.controller.sequence_finished.connect(self._on_sequence_finished_ui)
        self.controller.error_occurred.connect(lambda msg: self.progress_panel.parser_log.error(msg))
        self.controller.ui_lock_requested.connect(self.controls_widget.set_ui_locked)
        self.controller.request_increment.connect(self._on_request_increment)
        self.controller.ai_progress_updated.connect(self._on_ai_progress)
        self.controller.ai_batch_finished.connect(self._on_ai_batch_finished)
        self.controller.scan_finished.connect(self._on_scan_finished)
        self.controller.ai_all_finished.connect(self._on_ai_all_finished)
        self.search_widget.scan_categories_requested.connect(lambda tags: self.controller.scan_categories(tags))
        self.search_widget.categories_selected.connect(self._on_categories_selected)
        self.search_widget.categories_changed.connect(self._update_cost_calculation)
        self.search_widget.apply_tags_to_new_queue_requested.connect(self.on_apply_tags_to_new_queue_requested)
        self.search_widget.parameters_update_requested.connect(self.controls_widget.set_parameters)
        self.controls_widget.start_requested.connect(self._on_start_search)
        self.controls_widget.stop_requested.connect(self._on_stop_search)
        self.controls_widget.parameters_changed.connect(self._on_parameters_changed)
        self.controls_widget.parameters_changed.connect(self._update_cost_calculation)
        if hasattr(self.controls_widget, 'queue_manager_widget'):
            self.controls_widget.queue_manager_widget.queue_changed.connect(self._on_queue_changed)
            self.controls_widget.queue_manager_widget.queue_removed.connect(self._on_queue_removed)
            self.controls_widget.queue_manager_widget.queue_toggled.connect(self._on_queue_toggled)
        self.results_area.file_loaded.connect(self._on_file_loaded)
        self.results_area.file_deleted.connect(self._on_file_deleted)
        self.results_area.table_item_deleted.connect(self._on_table_item_deleted)
        self.results_area.table_closed.connect(self._on_table_closed)
        self.results_area.item_starred.connect(self._on_item_starred)
        self.results_area.analyze_table_requested.connect(self.on_analyze_table_requested)
        self.results_area.add_to_memory_requested.connect(self.on_add_to_memory_requested)
        self.results_area.export_table_requested.connect(self.on_export_table_requested)
        self.results_area.mini_browser.analyze_file_requested.connect(self.on_analyze_file_requested)
        self.results_area.mini_browser.addmemory_file_requested.connect(self.on_addmemory_file_requested)
        self.results_area.mini_browser.export_file_requested.connect(self.on_export_file_requested)
        self.results_area.results_table.analyze_item_requested.connect(self.on_analyze_item_requested)
        self.results_area.results_table.addmemory_item_requested.connect(self.on_addmemory_item_requested)
        self.controls_widget.pause_neuronet_requested.connect(self._on_pause_neuronet_requested)
        self.analytics_widget.send_message_signal.connect(self.on_chat_message_sent)
        self.controller.ai_chat_reply.connect(self.analytics_widget.on_ai_reply)
        self.controller.ai_result_ready.connect(self.handle_ai_result)

    def _on_sequence_finished_ui(self):
        self.queue_manager.clear_scanned_categories_bulk()

        if hasattr(self, 'search_widget'):
            self.search_widget.set_scanned_categories([])
            self.search_widget.set_forced_categories([])

        self.progress_panel.set_finished_state()
        self.controls_widget.set_ui_locked(False)

    def _update_cost_calculation(self):
        category_count = self.search_widget.get_category_count()
        category_count = max(1, category_count)
        self.controls_widget.update_category_count(category_count)

    def _on_table_closed(self):
        logger.info("Таблица закрыта пользователем...")

    def _save_current_queue_state(self):
        if getattr(self, '_is_programmatic_update', False):
            return
        
        idx = self.queue_manager.get_current_index()
        state = self.controls_widget.get_parameters()
        state.update({
            "search_tags": self.search_widget.get_search_tags(),
            "ignore_tags": self.search_widget.get_ignore_tags(),
            "forced_categories": self.search_widget.get_forced_categories(),
            "scanned_categories": self.search_widget.get_scanned_categories()
        })
        if hasattr(self.controls_widget, "queue_manager_widget"):
             current_name = self.controls_widget.queue_manager_widget.get_queue_name(idx)
             state["name"] = current_name
        self.queue_manager.set_state(state, idx)
    
    def _sync_queues_with_ui(self):
        if not hasattr(self.controls_widget, "queue_manager_widget"):
            return

        ui_mgr = self.controls_widget.queue_manager_widget
        ui_count = ui_mgr.get_all_queues_count()

        for idx in range(ui_count):
            self.queue_manager.get_state(idx)

        for idx in list(self.queue_manager.get_all_queue_indices()):
            if idx >= ui_count:
                self.queue_manager.delete_queue(idx)

    def _restore_queues_from_state(self):
        if not hasattr(self.controls_widget, "queue_manager_widget"):
            return

        indices = self.queue_manager.get_all_queue_indices()
        if not indices:
            return

        ui_mgr = self.controls_widget.queue_manager_widget

        while ui_mgr.get_all_queues_count() < len(indices):
            ui_mgr.add_queue()

        for idx in indices:
            state = self.queue_manager.get_state(idx)
            is_enabled = state.get("queue_enabled", True)
            ui_mgr.set_queue_checked(idx, is_enabled)

            saved_name = state.get("name", "")
            if saved_name:
                ui_mgr.set_queue_name(idx, saved_name)

    def _load_queue_to_ui(self, index: int):
        self._is_programmatic_update = True
        try:
            self.queue_manager.set_current_index(index)
            state = self.queue_manager.get_state(index)

            self.search_widget.set_search_tags(state.get("search_tags", []))
            self.search_widget.set_ignore_tags(state.get("ignore_tags", []))

            scanned_cats = state.get("scanned_categories", [])
            if scanned_cats:
                self.search_widget.set_scanned_categories(scanned_cats)

            forced_cats = state.get("forced_categories", [])
            self.search_widget.set_forced_categories(forced_cats)

            self.controls_widget.set_parameters(state)
            self._update_cost_calculation()

            is_enabled = state.get("queue_enabled", True)
            if hasattr(self.controls_widget, 'queue_manager_widget'):
                self.controls_widget.queue_manager_widget.set_queue_checked(index, is_enabled)

            logger.info(f"Загружена очередь #{index + 1}...")
            
        finally:
            self._is_programmatic_update = False

    def _safe_switch_queue_ui(self, target_index: int):
        self.queue_manager.set_current_index(target_index)

        if hasattr(self.controls_widget, 'queue_manager_widget'):
            ui_mgr = self.controls_widget.queue_manager_widget
            was_blocked = ui_mgr.list_widget.signalsBlocked()
            ui_mgr.list_widget.blockSignals(True)
            try:
                ui_mgr.set_current_queue(target_index)
            finally:
                ui_mgr.list_widget.blockSignals(was_blocked)

        self._load_queue_to_ui(target_index)

    def _on_queue_changed(self, new_index: int):
        self._save_current_queue_state()
        self._load_queue_to_ui(new_index)

    def _on_queue_toggled(self, index: int, is_checked: bool):
        self.queue_manager.update_state({"queue_enabled": is_checked}, index)
        
        status = "включена" if is_checked else "отключена"
        logger.info(f"Очередь #{index + 1} {status}...")

    def _on_queue_renamed(self, index: int, new_name: str):
        self.queue_manager.update_state({"name": new_name}, index)

    def _on_queue_removed(self, index: int):
        self.queue_manager.delete_queue(index)
        
        new_ui_index = index
        if hasattr(self.controls_widget, "queue_manager_widget"):
            count = self.controls_widget.queue_manager_widget.get_all_queues_count()
            if new_ui_index >= count:
                new_ui_index = count - 1
        
        if new_ui_index >= 0:
            self.queue_manager.set_current_index(new_ui_index)
            self._load_queue_to_ui(new_ui_index)

    def _on_pause_neuronet_requested(self):
        self.progress_panel.ai_log.warning(f"Функция паузы нейросети еще не реализована в контроллере.")

    def _on_start_search(self):
        self._save_current_queue_state()
        
        all_indices = self.queue_manager.get_all_queue_indices()
        self._validation_queue = []
        
        for idx in all_indices:
            state = self.queue_manager.get_state(idx)
            if state.get("queue_enabled", True):
                tags = state.get("search_tags", [])
                forced = state.get("forced_categories", [])
                
                if tags and not forced:
                    self._validation_queue.append(idx)

        if self._validation_queue:
            self.controls_widget.set_ui_locked(True)
            self._process_next_validation_step()
        else:
            self._finalize_and_start_search()

    def _process_next_validation_step(self):
        if not self._validation_queue:
            self._validating_index = None
            self._finalize_and_start_search()
            return

        idx = self._validation_queue.pop(0)
        self._validating_index = idx
        
        queue_name = f"#{idx + 1}"
        state = self.queue_manager.get_state(idx)

        custom_name = state.get("name", "")
        if custom_name:
             queue_name = f"{custom_name}"
        else:
            tags = state.get("search_tags", [])
            if tags:
                tags_str = ", ".join(tags[:2])
                if len(tags) > 2: tags_str += "..."
                queue_name += f" ({tags_str})"

        dlg = ScanPromptDialog(queue_name, self)
        res = dlg.exec()
        
        if res == 10:
            logger.info(f"Запуск сканирования категорий для очереди {queue_name}...")
            tags = state.get("search_tags", [])
            self.controller.scan_categories(tags)
            
        elif res == 20:
            logger.info(f"Очередь {queue_name}: выбран авто-режим категорий...")
            self._validating_index = None
            self._process_next_validation_step()
            
        else:
            logger.info("Запуск отменен пользователем...")
            self._validation_queue.clear()
            self._validating_index = None
            self.controls_widget.set_ui_locked(False)

    def _get_current_ai_instructions(self) -> str:
        if hasattr(self, 'memory_panel'):
            instr_list = self.memory_panel.get_instructions()
            return "\n".join([f"- {i}" for i in instr_list])
        return ""

    def _finalize_and_start_search(self):
        all_indices = self.queue_manager.get_all_queue_indices()
        active_configs = []
        user_instr = self._get_current_ai_instructions()
        
        ui_mgr = None
        if hasattr(self.controls_widget, "queue_manager_widget"):
            ui_mgr = self.controls_widget.queue_manager_widget

        for idx in all_indices:
            state = self.queue_manager.get_state(idx)
            if state.get("queue_enabled", True):
                if not state.get("search_tags"):
                    logger.warning(f"Очередь #{idx+1} пропущена: нет тегов...")
                    continue
                state['original_index'] = idx
                if ui_mgr:
                    q_name = ui_mgr.get_queue_name(idx)
                    state['queue_name'] = q_name
                active_configs.append(state)
        
        if not active_configs:
            self.controls_widget.set_ui_locked(False)
            QMessageBox.warning(self, "Ошибка", "Нет активных очередей с тегами для поиска!")
            return

        # --- ПОЛУЧЕНИЕ ГЛОБАЛЬНОГО РЕЖИМА РАЗДЕЛЕНИЯ ---
        global_split_mode = False
        if hasattr(self.controls_widget, "split_results_sw"):
            global_split_mode = self.controls_widget.split_results_sw.isChecked()

        # Принудительно применяем глобальный режим ко всем очередям
        for cfg in active_configs:
            cfg['split_results'] = global_split_mode

        self.is_sequence_running = True
        self.controls_widget.set_ui_locked(True)
        
        self.current_search_mode = active_configs[0].get("search_mode", "full")
        self.progress_panel.set_parser_mode(self.current_search_mode)
        self.cnt_parser = 0
        self.cnt_neuro = 0
        
        first_config = active_configs[0]
        
        # Если разделение ВЫКЛЮЧЕНО (глобально) - готовим один общий файл
        if not global_split_mode:
            merge_file = first_config.get("merge_with_table")
            if merge_file and os.path.exists(merge_file):
                self.current_json_file = merge_file
                self.current_results = self._load_results_file_silent(merge_file)
                logger.info(f"Режим добавления: {os.path.basename(merge_file)}...")
            else:
                first_queue_name = first_config.get('queue_name', '')
                self._create_new_results_file(queue_name=first_queue_name)
                self.current_results = []
                logger.info(f"Режим слияния: создан новый файл для всех очередей...")
        
        base_count = len(self.current_results or [])
        
        for cfg in active_configs:
            cfg['debug_mode'] = self.app_settings.get('debug_mode', False)
            cfg['ai_debug_mode'] = self.app_settings.get('ai_debug', False)
            cfg['user_instructions'] = user_instr
            
            # Назначаем целевую таблицу в зависимости от глобального режима
            if not global_split_mode:
                cfg['context_table'] = self.current_json_file
            else:
                # В режиме сплита проверяем персональную таблицу очереди
                custom_table = cfg.get('merge_with_table')
                if custom_table and os.path.exists(custom_table):
                    cfg['context_table'] = custom_table
                else:
                    cfg['context_table'] = None

            cfg['ai_offset'] = base_count
        
        if active_configs:
            first_idx = active_configs[0]['original_index']
            self._safe_switch_queue_ui(first_idx)

        self.progress_panel.set_parser_mode(self.current_search_mode)
        self.controller.start_sequence(active_configs)
        logger.info(f"Запуск {len(active_configs)} очередей...")

    def _on_stop_search(self):
        self.controller.request_soft_stop()
        logger.info(f"Остановка по запросу пользователя...")
        
        if hasattr(self.controls_widget, 'btn_stop'):
            self.controls_widget.stop_button.setEnabled(False)
            self.controls_widget.stop_button.setText("Останавливаемся...")

        self.is_sequence_running = False
        if self.controller.queue_state:
            self.controller.queue_state.is_sequence_running = False

        QTimer.singleShot(2000, self._force_ui_reset)

    def _force_ui_reset(self):
        self.controls_widget.set_ui_locked(False)
        self.progress_panel.parser_bar.setValue(0)
        self.progress_panel.reset_parser_progress() 
        
        if hasattr(self.controls_widget, 'btn_stop'):
            self.controls_widget.stop_button.setText("Остановить")

    def _on_parameters_changed(self, params: dict):
        merge_target = params.get("merge_with_table")
        
        is_split = params.get("split_results", False)
        has_context = bool(merge_target and not is_split)
        
        self.controls_widget.set_rewrite_controls_enabled(has_context)

    def _merge_results(self, new_items: List[Dict], base_items: List[Dict], rewrite_duplicates: bool):
        base_map = {str(item.get("id", "")): item for item in base_items if item.get("id")}
        added = 0
        updated = 0
        skipped = 0
        processed_ids = set()
        new_entries = []
        
        for item in new_items:
            iid = str(item.get("id", ""))
            if not iid: continue
            
            if iid in base_map:
                if rewrite_duplicates:
                    old = base_map[iid]
                    user_flags = {
                        "starred": old.get("starred", False),
                        "ai_comment": old.get("ai_comment", "") if not item.get("ai_comment") else item.get("ai_comment")
                    }
                    base_map[iid].update(item)
                    base_map[iid].update(user_flags)
                    updated += 1
                else: 
                    skipped += 1
            else:
                if iid not in processed_ids:
                    item.setdefault("starred", False)
                    new_entries.append(item)
                    added += 1
            processed_ids.add(iid)
            
        final_list = list(base_map.values()) + new_entries
        return final_list, added, updated, skipped

    def _sanitize_filename(self, name: str) -> str:
        return "".join(c if c.isalnum() or c in (' ', '-', '_', '.') else '_' for c in name).strip()

    def _on_parsing_finished(self, results: List[Dict]):
        idx = self.controller.queue_state.current_queue_index
        if idx < len(self.controller.queue_state.queues_config):
            config = self.controller.queue_state.queues_config[idx]
        else:
            config = self.controls_widget.get_parameters()

        search_mode = config.get("search_mode", "full")

        if search_mode == "neuro":
            logger.info(
                f"Очередь #{idx + 1}: парсинг завершен, ожидаем нейро-фильтр "
                f"({len(results)} сырых элементов)..."
            )
            return

        self._save_results_for_config(results, config, idx)

        if not self.controller.queue_state.is_sequence_running:
            self.controls_widget.set_ui_locked(False)
            self.is_sequence_running = False
            self.progress_panel.set_finished_state()
            logger.success("Парсинг завершен...")

    def _create_new_results_file(self, queue_name: str = ""):
        timestamp = time.strftime('%d%m%Y_%H%M%S')
        
        if queue_name:
            safe_name = self._sanitize_filename(queue_name)
            filename = f"avito_{safe_name}_{timestamp}.json"
        else:
            filename = f"avito_search_{timestamp}.json"
            
        self.current_json_file = os.path.join(RESULTS_DIR, filename)

    def _save_results_to_file(self):
        if not self.current_json_file:
            return
        self._save_list_to_file(self.current_results, self.current_json_file)

    def _save_list_to_file(self, data: list, filepath: str):
        import gzip
        try:
            os.makedirs(os.path.dirname(filepath), exist_ok=True)
            with gzip.open(filepath, 'wt', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            self._refresh_merge_targets()
        except Exception as e:
            pass

    def _on_parser_started_logic(self):
        qs = self.controller.queue_state
        idx = qs.current_queue_index

        if 0 <= idx < len(qs.queues_config):
            self.current_search_mode = qs.queues_config[idx].get("search_mode", "full")

        self.progress_panel.set_parser_mode(self.current_search_mode)

        if self.current_search_mode != "primary":
            self.progress_panel.reset_parser_progress()

        self.progress_panel.reset_ai_progress()
        
    def _on_queue_finished(self, results, idx, is_split):
        config = {}
        if idx < len(self.controller.queue_state.queues_config):
            config = self.controller.queue_state.queues_config[idx]
        q_name = config.get('queue_name', f"#{idx + 1}")
        logger.success(f"#{q_name}: поиск завершен. Получено {len(results)} шт...")
        
        next_idx = idx + 1
        if next_idx < len(self.controller.queue_state.queues_config):
            next_config = self.controller.queue_state.queues_config[next_idx]
            original_next_idx = next_config.get('original_index', next_idx)
        
            next_mode = next_config.get("search_mode", "full")
            self.current_search_mode = next_mode
            self.progress_panel.set_parser_mode(next_mode)
        
            if next_mode != "primary":
                self.progress_panel.reset_parser_progress()
        
            self._safe_switch_queue_ui(original_next_idx)

    def _on_parser_progress(self, val: int):
        clamped_val = max(0, min(100, val))
        
        if self.current_search_mode != "primary":
            if self.progress_panel.parser_bar.maximum() == 0:
                 self.progress_panel.parser_bar.setRange(0, 100)
            
            self.progress_panel.parser_bar.setValue(clamped_val)
        
        if clamped_val > 0 and clamped_val < 100:
            logger.progress(f"Выполнение: {clamped_val}%", token="parser_global_progress")
        elif clamped_val == 100:
            logger.success("Сбор данных завершен (100%)") # TODO: 1. Это вызывается три раза, если верить Журналу событий? 2. Перенести это в парсер?
             
        self.last_progress_value = clamped_val

    def _check_ai_availability(self):
        self.controller.ensure_ai_manager()
        
        if not self.controller.ai_manager.has_model():
            self._disable_ai_features()
            self._show_no_model_notification()
        else:
            self._enable_ai_features()

    # TODO
    def _show_no_model_notification(self):
        msg_box = QMessageBox(self)
        msg_box.setIcon(QMessageBox.Icon.Information)
        msg_box.setWindowTitle("AI модель не найдена")
        msg_box.setText(
            "<b>Модель нейросети не обнаружена</b><br><br>"
            "Все ИИ-функции будут недоступны до установки модели:<br>"
            "• Вкладка \"Аналитика\"<br>"
            "• Режим поиска \"Нейро\"<br>"
            "• Опции \"Вкл. в анализ\" и \"RAG\"<br><br>"
            "Откройте <b>Настройки → Нейросеть</b> для скачивания стандартной модели."
        )
        msg_box.setStandardButtons(QMessageBox.StandardButton.Ok)
        msg_box.setStyleSheet(Components.dialog())
        msg_box.exec()

    def _enable_ai_features(self):
        self.btn_analytics.setEnabled(True)
        self.btn_analytics.setToolTip("")
        self.btn_analytics.setStyleSheet(Components.nav_button())
        
        if hasattr(self.search_widget, 'search_mode_widget'):
            mode_widget = self.search_widget.search_mode_widget
            if hasattr(mode_widget, 'btn_neuro'):
                mode_widget.btn_neuro.setEnabled(True)
                mode_widget.btn_neuro.setToolTip("Анализ с ИИ")
                mode_widget._update_buttons()
        
        if hasattr(self.controls_widget, 'include_ai_sw'):
            self.controls_widget.include_ai_sw.setEnabled(True)
            self.controls_widget.include_ai_sw.setToolTip("")

            if hasattr(self.controls_widget, 'include_ai_lbl'):
                is_on = self.controls_widget.include_ai_sw.isChecked()
                self.controls_widget.include_ai_lbl.setText("Вкл" if is_on else "Выкл")
                self.controls_widget.include_ai_lbl.setStyleSheet(
                    Typography.style(
                        family=Typography.UI,
                        size=Typography.SIZE_SMALL,
                        weight=Typography.WEIGHT_BOLD if is_on else Typography.WEIGHT_NORMAL,
                        color=Palette.SECONDARY if is_on else Palette.TEXT_MUTED
                    )
                )
        
        if hasattr(self.controls_widget, 'store_memory_sw'):
            self.controls_widget.store_memory_sw.setEnabled(True)
            self.controls_widget.store_memory_sw.setToolTip("")
            self.controls_widget.store_memory_sw.setEnabled(self.controller.ai_manager.has_model()) 

            if hasattr(self.controls_widget, 'store_memory_lbl'):
                is_on = self.controls_widget.store_memory_sw.isChecked()
                self.controls_widget.store_memory_lbl.setText("Вкл" if is_on else "Выкл")
                self.controls_widget.store_memory_lbl.setStyleSheet(
                    Typography.style(
                        family=Typography.UI,
                        size=Typography.SIZE_SMALL,
                        weight=Typography.WEIGHT_BOLD if is_on else Typography.WEIGHT_NORMAL,
                        color=Palette.SECONDARY if is_on else Palette.TEXT_MUTED
                    )
                )
        
        if hasattr(self.controls_widget, 'ai_criteria_container'):
            mode = self.controls_widget.search_mode_widget.get_mode()
            self.controls_widget.ai_criteria_container.setEnabled(mode == "neuro")

            if hasattr(self.controls_widget, 'ai_criteria_input'):
                self.controls_widget.ai_criteria_input.setPlaceholderText(
                    "Например: только на гарантии, с оригинальной упаковкой..."
                )
                self.controls_widget.ai_criteria_input.setStyleSheet(
                    Components.text_input()
                )
        
        logger.success(f"ИИ функции доступны...")

        if hasattr(self, '_model_was_just_downloaded') and self._model_was_just_downloaded:
            self._model_was_just_downloaded = False
            QMessageBox.information(
                self,
                "ИИ активирован",
                "Модель успешно загружена!\n"
                "Все ИИ-функции теперь доступны:\n\n"
                "• Вкладка Аналитика\n"
                "• Режим поиска Нейро\n"
                "• Опции \"Вкл. в анализ\" и \"RAG\"",
                QMessageBox.StandardButton.Ok
            )

    def _disable_ai_features(self):
        self.btn_analytics.setEnabled(False)
        self.btn_analytics.setToolTip("Требуется модель нейросети. Откройте 'Настройки'.")
        
        self.btn_analytics.setStyleSheet(f"""
            QPushButton {{
                background-color: {Palette.BG_DARK_3};
                color: {Palette.TEXT_MUTED};
                border: 1px solid {Palette.BORDER_SOFT};
                border-radius: {Spacing.RADIUS_NORMAL}px;
                padding: {Spacing.SM}px {Spacing.MD}px;
                opacity: 0.5;
            }}
            QPushButton:hover {{
                border-color: {Palette.WARNING};
                color: {Palette.WARNING};
            }}
        """)

        if hasattr(self.search_widget, 'search_mode_widget'):
            mode_widget = self.search_widget.search_mode_widget
            if hasattr(mode_widget, 'btn_neuro'):
                mode_widget.btn_neuro.setEnabled(False)
                mode_widget.btn_neuro.setToolTip(
                    "Нейро-режим недоступен\n"
                    "Требуется модель AI\n\n"
                    "Откройте Настройки → Нейросеть для установки"
                )

                mode_widget.btn_neuro.setStyleSheet(f"""
                    QPushButton {{
                        background-color: {Palette.BG_DARK_3};
                        border: 1px solid {Palette.BORDER_SOFT};
                        border-radius: {Spacing.RADIUS_NORMAL}px;
                        color: {Palette.TEXT_MUTED};
                        padding: {Spacing.XS}px {Spacing.MD}px;
                        text-decoration: line-through;
                    }}
                    QPushButton:hover {{
                        border-color: {Palette.WARNING};
                        background-color: {Palette.with_alpha(Palette.WARNING, 0.1)};
                    }}
                """)

                current_mode = mode_widget.get_mode()
                if current_mode == "neuro":
                    mode_widget.set_mode("full")

        if hasattr(self.controls_widget, 'include_ai_sw'):
            self.controls_widget.include_ai_sw.setEnabled(False)
            self.controls_widget.include_ai_sw.setChecked(False)
            self.controls_widget.include_ai_sw.setToolTip(
                "Требуется модель нейросети\nОткройте Настройки"
            )

            if hasattr(self.controls_widget, 'include_ai_lbl'):
                self.controls_widget.include_ai_lbl.setText("Недоступно")
                self.controls_widget.include_ai_lbl.setStyleSheet(
                    Typography.style(
                        family=Typography.UI,
                        size=Typography.SIZE_SMALL,
                        color=Palette.TEXT_MUTED
                    )
                )

        if hasattr(self.controls_widget, 'store_memory_sw'):
            self.controls_widget.store_memory_sw.setEnabled(False)
            self.controls_widget.store_memory_sw.setChecked(False)
            self.controls_widget.store_memory_sw.setToolTip(
                "RAG требует модель нейросети"
            )

            if hasattr(self.controls_widget, 'store_memory_lbl'):
                self.controls_widget.store_memory_lbl.setText("Недоступно")
                self.controls_widget.store_memory_lbl.setStyleSheet(
                    Typography.style(
                        family=Typography.UI,
                        size=Typography.SIZE_SMALL,
                        color=Palette.TEXT_MUTED
                    )
                )
        
        if hasattr(self.controls_widget, 'ai_criteria_container'):
            self.controls_widget.ai_criteria_container.setEnabled(False)

            if hasattr(self.controls_widget, 'ai_criteria_input'):
                self.controls_widget.ai_criteria_input.setPlaceholderText(
                    "Недоступно: требуется модель нейросети"
                )
                self.controls_widget.ai_criteria_input.setStyleSheet(
                    Components.text_input() + f"""
                    QPlainTextEdit {{
                        background-color: {Palette.BG_DARK_2};
                        color: {Palette.TEXT_MUTED};
                        opacity: 0.5;
                    }}
                    """
                )
        
        logger.warning(f"ИИ функции отключены: модель не найдена...")
        logger.info(f"Откройте 'Настройки → Нейросеть' для установки модели...")

    def handle_ai_result(self, idx, json_text, context):
        self.results_area.results_table.update_ai_column(idx, json_text)

        if context.get('storeinmemory'):
            items = context.get('items', [])
            if idx < len(items):
                item = items[idx]
                parsed = json.loads(json_text)
                item['verdict'] = parsed.get('verdict')
                item['reason'] = parsed.get('reason')
                item['market_position'] = parsed.get('market_position')
                item['defects'] = parsed.get('defects')
                self.memory_manager.add_item(item)

    def on_chat_message_sent(self, chat_history: list):
        if not self.controller.ai_manager:
            self.analytics_widget.on_ai_reply("ИИ не доступен. Загрузите модель в настройках.")
            return

        user_instructions = []
        if hasattr(self, 'memory_panel'):
            user_instructions = self.memory_panel.get_instructions()

        current_table = self.current_results if self.current_results else []

        self.controller.ai_manager.start_chat_request(
            chat_history, 
            user_instructions=user_instructions,
            current_table_data=current_table
        )

    def rebuild_rag_cache(self):
        if self.is_sequence_running:
            return

        import threading
        def rebuild_task():
            try:
                if hasattr(self, "analyticswidget"):
                    self.analytics_widget.refresh_data()
            except Exception as e:
                pass

        threading.Thread(target=rebuild_task, daemon=True).start()

    def _on_settings_closed_with_model(self):
        if self.controller.ai_manager and self.controller.ai_manager.has_model():
            self._enable_ai_features()
            QMessageBox.information(
                self,
                "ИИ активирован",
                "Модель успешно загружена!\nВсе ИИ-функции теперь доступны.",
                QMessageBox.StandardButton.Ok
            )

    def _on_ai_progress(self, val: int):
        self.progress_panel.ai_bar.setValue(val)

    def _on_ai_result(self, idx, json_text, ctx):
        if 0 <= idx < len(self.current_results):
            self.current_results[idx]["ai_comment"] = json_text
            self.results_area.results_table.update_ai_column(idx, json_text)
            self._save_results_to_file()

    def _on_ai_all_finished(self):
        logger.info(f"ИИ завершил анализ...")

    def _on_ai_result_with_memory(self, index: int, json_text: str, context: dict):
        from app.core.log_manager import logger
        import json

        mode = context.get('mode', 'analysis')
        queue_idx = context.get('queue_idx', -1)
        if queue_idx is None: queue_idx = context.get('queueidx', 0)
        
        store_in_memory = context.get('store_in_memory', False)
        items = context.get('items') or []
        
        try:
            data = json.loads(json_text) if isinstance(json_text, str) else json_text
        except Exception as e:
            logger.dev(f"AI JSON parse error: {e} | Payload: {str(json_text)[:200]}", level="ERROR")
            return

        src_item = None
        if isinstance(items, list) and 0 <= index < len(items):
            src_item = items[index]

        if mode == 'filter':
            verdict = str(data.get("verdict", "")).upper()
            reason = data.get("reason", "")
            
            if not src_item: return
            
            src_item['ai_filter_verdict'] = verdict
            src_item['ai_filter_reason'] = reason
            
            if verdict == "GOOD":
                bucket = self._neuro_filtered.setdefault(queue_idx, [])
                bucket.append(src_item)
            return

        if mode == 'analysis':
            if not src_item: return
            
            src_item["ai"] = data
            src_item["verdict"] = data.get("verdict")
            src_item["reason"] = data.get("reason")
            src_item["market_position"] = data.get("market_position")
            src_item["defects"] = data.get("defects")

            if queue_idx == -1: 
                self.results_area.results_table.update_ai_column(index, json_text)
            
            if self.current_results:
                item_id = str(src_item.get('id'))
                found = False
                for existing in self.current_results:
                    if str(existing.get('id')) == item_id:
                        existing.update(src_item)
                        found = True
                        break
                
                if found and self.current_json_file:
                     self._save_results_to_file()

            if store_in_memory:
                try:
                    self.memory_manager.add_item(src_item)
                except Exception as e:
                    logger.dev(f"Memory add_item error: {e}", level="ERROR")

    def _save_results_for_config(self, results: List[Dict], config: Dict, idx: int):
        q_name = config.get("queue_name", f"Queue {idx+1}")
        if not results:
            logger.warning(f"Очередь '{q_name}': результатов нет, файл не создан...")
            return
        
        split_results = config.get("split_results", False)
        rewrite = config.get("rewrite_duplicates", False)

        if split_results:
            target_file = config.get('context_table')
            
            if target_file and os.path.exists(target_file):
                base_items = self._load_results_file_silent(target_file)
                merged, added, updated, skipped = self._merge_results(results, base_items, rewrite)
                self._save_list_to_file(merged, target_file)
                
                if self.current_json_file == target_file:
                     self.current_results = merged
                     self.results_area.load_full_history(merged)
                
                logger.success(f"Очередь '{q_name}': Добавлено {added}, Обновлено {updated} в {os.path.basename(target_file)}")
                
                if hasattr(self.results_area, "mini_browser"):
                    self.results_area.mini_browser.refresh_files()

            else:
                timestamp = time.strftime('%d%m%Y_%H%M%S')
                raw_queue_name = config.get('queue_name', '')
                if not raw_queue_name:
                    q_num = config.get('original_index', idx) + 1
                    raw_queue_name = f"queue_{q_num}"
                
                safe_name = self._sanitize_filename(raw_queue_name)
                filename = f"avito_{safe_name}_{timestamp}.json"
                target_file = os.path.join(RESULTS_DIR, filename)

                base_items: List[Dict] = []
                merged, added, updated, skipped = self._merge_results(results, base_items, rewrite)
                self._save_list_to_file(merged, target_file)
                
                logger.success(f"Очередь '{raw_queue_name}': Сохранено в {filename} (добавлено {added})...")

                if hasattr(self.results_area, "mini_browser"):
                    self.results_area.mini_browser.refresh_files()

        else:
            if not self.current_json_file:
                 self._create_new_results_file(queue_name=q_name)

            base_items = self.current_results
            merged, added, updated, skipped = self._merge_results(results, base_items, rewrite)
            self.current_results = merged
            self._save_results_to_file()

            self.results_area.load_full_history(self.current_results)

            if self.current_json_file:
                basename = os.path.basename(self.current_json_file).replace("avito_", "").replace(".json", "")
                fulldate = time.strftime("%d.%m.%Y %H:%M")
                self.results_area.update_header(
                    table_name=basename,
                    full_date=fulldate,
                    count=len(self.current_results)
                )
                if hasattr(self.results_area, "mini_browser"):
                    self.results_area.mini_browser.refresh_files()

    def _open_rag_tab(self):
        self._switch_page(1)
        if hasattr(self.analytics_widget, "tabs"):
            self.analytics_widget.tabs.setCurrentIndex(0)

    def _on_scan_finished(self, categories):
        if self._validating_index is not None:
            logger.info(f"Категории получены. Открываем выбор для очереди #{self._validating_index + 1}...")
            
            state = self.queue_manager.get_state(self._validating_index)
            current_forced = state.get("forced_categories", [])
            
            def on_clear():
                 self.queue_manager.update_state({"forced_categories": []}, self._validating_index)

            dlg = CategorySelectionDialog(categories, self, current_forced, on_clear)
            if dlg.exec():
                selected = dlg.get_selected()
                self.queue_manager.update_state({
                    "forced_categories": selected,
                    "scanned_categories": categories
                }, self._validating_index)
                logger.info(f"Для очереди #{self._validating_index + 1} сохранено {len(selected)} категорий.")
                
                if self.queue_manager.get_current_index() == self._validating_index:
                     self.search_widget.set_forced_categories(selected)
                     self._update_cost_calculation()

                self._validating_index = None
                self._process_next_validation_step()
            else:
                logger.info("Выбор категорий отменен. Запуск остановлен.")
                self._validation_queue.clear()
                self._validating_index = None
                self.controls_widget.set_ui_locked(False)
            return

        self.search_widget.set_scanned_categories(categories)
        logger.info(f"Найдено {len(categories)} категорий...")

    def _on_categories_selected(self, cats):
        logger.info(f"Выбрано {len(cats)} категорий...")

    def on_apply_tags_to_new_queue_requested(self, search_tags: list, ignore_tags: list, tag_params: dict = None):
        self._save_current_queue_state()
        
        if not hasattr(self.controls_widget, "params_panel"):
            return
        
        ui_mgr = self.controls_widget.params_panel.queue_manager_widget
        new_index = ui_mgr.get_all_queues_count()
        ui_mgr.add_queue()
        
        # 1. Создаем ЧИСТОЕ состояние (заводские настройки)
        base_state = self.queue_manager._create_default_state()
        
        # ВАЖНО: Мы НЕ делаем base_state.update(current_active), 
        # чтобы не тащить настройки из предыдущей очереди
        
        # 2. Ищем параметры тега
        last_p = None
        if tag_params:
            for t in search_tags:
                if t in tag_params:
                    last_p = tag_params[t]
                    # Если нашли параметры хотя бы для одного тега - используем их
                    break 
                
        # 3. Применяем параметры тега ТОЛЬКО если они есть (зеленый индикатор)
        if last_p:
            base_state.update(last_p)
        
        # 4. Устанавливаем общие поля
        base_state["search_tags"] = list(search_tags)
        base_state["ignore_tags"] = list(ignore_tags)
        base_state["forced_categories"] = []
        base_state["name"] = ""
        
        # 5. Применяем и сохраняем
        self.queue_manager.set_state(base_state, new_index)
        self.queue_manager.set_current_index(new_index)
        self.queue_manager.save_current_state()
        
        ui_mgr.list_widget.blockSignals(True)
        ui_mgr.set_current_queue(new_index)
        ui_mgr.list_widget.blockSignals(False)
        
        self._load_queue_to_ui(new_index)

    def _on_file_loaded(self, path, data):
        self.current_json_file = path 
        self.current_results = data
        self.results_area.load_full_history(data)

        import os
        from datetime import datetime
        basename = os.path.basename(path).replace("avito_", "").replace(".json", "")
        if os.path.exists(path):
            fulldate = datetime.fromtimestamp(os.path.getmtime(path)).strftime("%d.%m.%Y %H:%M")
        else:
            fulldate = datetime.now().strftime("%d.%m.%Y %H:%M")
            
        self.results_area.update_header(
            table_name=basename,
            full_date=fulldate,
            count=len(data)
        )

        if hasattr(self, 'tracker') and path:
            self.tracker.update_items_from_current_table(self.current_results, path)

        self._refresh_merge_targets()

    def _on_file_deleted(self, path):
        if self.current_json_file == path:
            self.current_json_file = None
            self.current_results = []
            self.results_area.clear_table()
        self._refresh_merge_targets()

    def on_analyze_table_requested(self, items: List[Dict]):
        if not items:
            QMessageBox.warning(self, "Ошибка", "Таблица пуста!")
            return

        user_criteria = ""

        logger.info(f"Запуск анализа для {len(items)} элементов...")
        
        self.controller.start_manual_ai_analysis(
            items=items,
            prompt=user_criteria, 
            debug_mode=self.app_settings.get("ai_debug", False),
            store_in_memory=self.controls_widget.get_parameters().get("store_in_memory", False)
        )

    def on_add_to_memory_requested(self, items: List[Dict]):
        """Добавить все элементы в RAG-память"""
        if not items:
            QMessageBox.warning(self, "Ошибка", "Таблица пуста!")
            return

        added = 0
        for item in items:
            if self.memory_manager.add_item(item):
                added += 1

        logger.success(f"Добавлено {added} из {len(items)} элементов в память ИИ")
        QMessageBox.information(self, "Успех", f"Добавлено {added} элементов в память ИИ")

    def on_export_table_requested(self, items: List[Dict]):
        """Экспортировать таблицу в файл"""
        if not items:
            QMessageBox.warning(self, "Ошибка", "Таблица пуста!")
            return

        from PyQt6.QtWidgets import QFileDialog
        import csv

        filepath, _ = QFileDialog.getSaveFileName(
            self, "Сохранить таблицу", "", "Excel файл (*.xlsx);;CSV файл (*.csv)"
        )

        if not filepath:
            return

        try:
            if filepath.endswith('.csv'):
                # Для CSV добавляем отдельную колонку со ссылкой
                fieldnames = ['id', 'price', 'title', 'link', 'views', 'date', 'city', 'description', 'ai_verdict']
                headers = ['ID', 'Цена', 'Название', 'Ссылка', 'Просмотров', 'Дата', 'Город', 'Описание', 'Вердикт ИИ']

                with open(filepath, 'w', encoding='utf-8-sig', newline='') as f:
                    writer = csv.writer(f)
                    writer.writerow(headers)
                    for item in items:
                        # Получить вердикт ИИ
                        ai_data = item.get('ai', {})
                        if isinstance(ai_data, str):
                            import json
                            try:
                                ai_data = json.loads(ai_data)
                            except:
                                ai_data = {}

                        verdict = ai_data.get('verdict', '')
                        reason = ai_data.get('reason', '')
                        ai_text = f"{verdict}: {reason}" if verdict else ""

                        desc = item.get('description', '') or ''
                        desc_short = desc[:100] + '...' if len(desc) > 100 else desc

                        writer.writerow([
                            item.get('id', ''),
                            item.get('price', ''),
                            item.get('title', ''),
                            item.get('link', ''),
                            item.get('views', ''),
                            item.get('date', ''),
                            item.get('address', ''),
                            desc_short,
                            ai_text
                        ])

            elif filepath.endswith('.xlsx'):
                try:
                    import openpyxl
                    from openpyxl.styles import Font
                except ImportError:
                    QMessageBox.critical(self, "Ошибка", "Для экспорта в Excel установите библиотеку: pip install openpyxl")
                    return

                headers = ['ID', 'Цена', 'Название (ссылка)', 'Просмотров', 'Дата', 'Город', 'Описание', 'Вердикт ИИ']

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(headers)

                for cell in ws[1]:
                    cell.font = Font(bold=True)

                for idx, item in enumerate(items, start=2):
                    ai_data = item.get('ai', {})
                    if isinstance(ai_data, str):
                        import json
                        try:
                            ai_data = json.loads(ai_data)
                        except:
                            ai_data = {}

                    verdict = ai_data.get('verdict', '')
                    reason = ai_data.get('reason', '')
                    ai_text = f"{verdict}: {reason}" if verdict else ""

                    desc = item.get('description', '') or ''
                    desc_short = desc[:100] + '...' if len(desc) > 100 else desc

                    title = item.get('title', '')
                    link = item.get('link', '')

                    ws.append([
                        item.get('id', ''),
                        item.get('price', ''),
                        title,
                        item.get('views', ''),
                        item.get('date', ''),
                        item.get('address', ''),
                        desc_short,
                        ai_text
                    ])

                    if link:
                        cell = ws.cell(row=idx, column=3)
                        cell.hyperlink = link
                        cell.font = Font(color="0563C1", underline="single")
                        cell.value = title

                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

                wb.save(filepath)

            logger.success(f"Таблица экспортирована: {filepath}")
            QMessageBox.information(self, "Успех", f"Таблица сохранена:\n{filepath}")
        except Exception as e:
            import traceback
            logger.error(f"Ошибка экспорта: {e}")
            traceback.print_exc()
            QMessageBox.critical(self, "Ошибка", f"Не удалось экспортировать:\n{e}")

    def on_analyze_file_requested(self, filepath: str):
        items = self._load_results_file_silent(filepath)
        if items:
            self.on_analyze_table_requested(items)

    def on_addmemory_file_requested(self, filepath: str):
        items = self._load_results_file_silent(filepath)
        if items:
            self.on_add_to_memory_requested(items)

    def on_export_file_requested(self, filepath: str):
        items = self._load_results_file_silent(filepath)
        if items:
            self.on_export_table_requested(items)

    def on_analyze_item_requested(self, item: Dict):
        user_criteria = ""
        
        logger.info(f"Запуск анализа для элемента {item.get('id', 'N/A')}...")
        
        self.controller.start_manual_ai_analysis(
            items=[item],
            prompt=user_criteria,
            debug_mode=self.app_settings.get("ai_debug", False),
            store_in_memory=self.controls_widget.get_parameters().get("store_in_memory", False)
        )

    def on_addmemory_item_requested(self, item: Dict):
        if self.memory_manager.add_item(item):
            logger.success(f"Элемент {item.get('id', 'N/A')} добавлен в память ИИ")
            QMessageBox.information(self, "Успех", "Элемент добавлен в память ИИ")
        else:
            logger.warning(f"Элемент {item.get('id', 'N/A')} уже в памяти")
            QMessageBox.information(self, "Информация", "Элемент уже есть в памяти")

    def _on_results_context_cleared(self):
        self.current_json_file = None
        self.current_results = []
        self.results_area.clear_table()
        self.controls_widget.set_rewrite_controls_enabled(False)
    
    def _on_result_selected(self):
        rows = self.results_area.results_table.selectedItems()
        if not rows:
            return

        row = rows[0].row()
        title_item = self.results_area.results_table.item(row, 2)
        if not title_item:
            return

    def _on_table_item_deleted(self, item_id):
        self.current_results = [x for x in self.current_results if str(x.get("id")) != str(item_id)]
        self._save_results_to_file()

    def _on_item_starred(self, item_id, is_starred):
        target_item = None
        for x in self.current_results:
            if str(x.get("id")) == str(item_id):
                x["starred"] = is_starred
                target_item = x
                break
        self._save_results_to_file()
        
        if hasattr(self, 'tracker'):
            if self.current_json_file:
                self.tracker.update_items_from_current_table(self.current_results, self.current_json_file)
            
            if is_starred and target_item:
                self.notifier.send_new_favorite(target_item)

    def _on_request_increment(self, p_count, ai_count):
        self.cnt_parser += p_count
        self.cnt_neuro += ai_count

    def _on_tracker_item_updated(self, updated_item):
        source_file = updated_item.get('_source_file')
        item_id = str(updated_item.get('id', ''))
        
        if self.current_json_file and source_file == self.current_json_file:
            for i, item in enumerate(self.current_results):
                if str(item.get('id', '')) == item_id:
                    self.current_results[i].update(updated_item)
                    break
            
            self._save_results_to_file()
            self.results_area.load_full_history(self.current_results)
            logger.info(f"Трекер обновил товар {item_id} (в текущем окне).")

        elif source_file and os.path.exists(source_file):
            try:
                import gzip
                data = []
                is_gzip = False
                try:
                    with open(source_file, 'r', encoding='utf-8') as f: data = json.load(f)
                except:
                    with gzip.open(source_file, 'rt', encoding='utf-8') as f: data = json.load(f)
                    is_gzip = True
                
                updated = False
                for i, item in enumerate(data):
                    if str(item.get('id', '')) == item_id:
                        data[i].update(updated_item)
                        if '_source_file' in data[i]: del data[i]['_source_file']
                        updated = True
                        break
                
                if updated:
                    if is_gzip:
                        with gzip.open(source_file, 'wt', encoding='utf-8') as f:
                            json.dump(data, f, ensure_ascii=False, indent=2)
                    else:
                        with open(source_file, 'w', encoding='utf-8') as f:
                            json.dump(data, f, ensure_ascii=False, indent=2)
                    logger.info(f"Трекер обновил товар {item_id} в фоновом файле: {os.path.basename(source_file)}")
                    
            except Exception as e:
                logger.error(f"Ошибка фонового обновления файла {source_file}: {e}")

    def _on_ai_batch_finished(self):
        self.progress_panel.ai_log.success("AI пакет обработан")

        qs = self.controller.queue_state
        if not qs.waiting_for_ai_sequence:
            return

        idx = qs.current_queue_index
        if idx >= len(qs.queues_config):
            return

        config = qs.queues_config[idx]
        search_mode = config.get("search_mode", "full")

        if search_mode != "neuro":
            return

        filtered = self._neuro_filtered.get(idx, []) or []
        logger.success(
            f"Нейро-фильтр для очереди #{idx + 1}: прошло {len(filtered)} объявлений."
        )

        self._neuro_filtered.pop(idx, None)

        self._save_results_for_config(filtered, config, idx)

        self.progress_panel.set_ai_finished()

        include_ai = config.get("include_ai", False)
        store_in_memory = config.get("store_in_memory", False)

        if filtered and (include_ai or store_in_memory):
            user_instructions = "" 
            ai_debug = config.get("ai_debug_mode", False)

            self.controller.start_manual_ai_analysis(
                items=filtered,
                prompt=user_instructions,
                debug_mode=ai_debug,
                store_in_memory=store_in_memory,
            )

    def _refresh_merge_targets(self):
        if hasattr(self.results_area, "mini_browser"):
            files = self.results_area.mini_browser.iter_files()
            targets = []
            for fpath, _, _ in files:
                name = os.path.basename(fpath).replace("avito_", "").replace(".json", "")
                targets.append((fpath, name))
            self.controls_widget.set_merge_targets(targets)
            has_context = bool(self.controls_widget.get_parameters().get("merge_with_table") or self.current_json_file)
            self.controls_widget.set_rewrite_controls_enabled(has_context)

    def _load_results_file_silent(self, path):
        if not path or not os.path.exists(path): return []
        import gzip
        try:
            try:
                with open(path, "r", encoding="utf-8") as f: return json.load(f)
            except:
                with gzip.open(path, "rt", encoding="utf-8") as f: return json.load(f)
        except: return []

    def _switch_page(self, index):
        
        self.stack.setCurrentIndex(index)
        
        self.btn_parser.setChecked(index == 0)
        self.btn_analytics.setChecked(index == 1)
        self.btn_memory.setChecked(index == 2)

    def _on_settings_clicked(self):
        dlg = SettingsDialog(self.app_settings, self)

        dlg.model_downloaded.connect(self._on_model_downloaded)
        dlg.factory_reset_requested.connect(self.close)

        if dlg.exec():
            self.app_settings = dlg.get_settings()
            self._save_settings()
            
            self.controller.ensure_ai_manager()
            if self.controller.ai_manager:
                self.controller.ai_manager.update_config(self.app_settings)
            
            if hasattr(self, 'tracker'):
                self.tracker.update_settings(self.app_settings)

    def _on_model_downloaded(self, file_path: str):
        self._model_was_just_downloaded = True

        if self.controller.ai_manager:
            model_name = os.path.basename(file_path)
            self.controller.ai_manager.set_model(model_name)

        self._enable_ai_features()

    def _load_settings(self):
        path = os.path.join(BASE_APP_DIR, "app_settings.json")
        settings = {}
        if os.path.exists(path):
            try:
                with open(path, "r", encoding="utf-8") as f: 
                    settings = json.load(f)
            except: pass
        
        model_path = settings.get("ai_model_path")
        if model_path:
            if not os.path.exists(model_path):
                from app.config import MODELS_DIR
                filename = os.path.basename(model_path)
                local_candidate = os.path.join(MODELS_DIR, filename)
                
                if os.path.exists(local_candidate):
                    logger.info(f"Путь к модели восстановлен локально: {local_candidate}")
                    settings["ai_model_path"] = local_candidate
                else:
                    logger.warning(f"Модель не найдена по пути: {model_path}. Сброс.")
                    settings["ai_model_path"] = ""
                    
        return settings
        
    def _save_settings(self):
        path = os.path.join(BASE_APP_DIR, "app_settings.json")
        try:
            with open(path, "w", encoding="utf-8") as f: json.dump(self.app_settings, f, indent=2)
        except: pass

    def _start_timers(self):
        self.ai_stats_timer = QTimer(self)
        self.ai_stats_timer.timeout.connect(self._update_ai_stats)
        self.ai_stats_timer.start(1000)

        self.rag_stats_timer = QTimer(self)
        self.rag_stats_timer.timeout.connect(self._update_rag_stats)
        self.rag_stats_timer.start(5000)

    def _update_ai_stats(self):
        if self.controller.ai_manager:
            stats = self.controller.ai_manager.refresh_resource_usage()
            # Check if panel is still there (it might be commented out)
            if hasattr(self, 'ai_stats_panel') and self.ai_stats_panel.isVisible():
                self.ai_stats_panel.update_stats(stats)

    def _update_rag_stats(self):
        stats = self.memory_manager.get_rag_status()
        if hasattr(self, 'rag_stats_panel') and self.rag_stats_panel.isVisible():
            self.rag_stats_panel.update_stats(stats)

    def apply_initial_geometry(self):
        app = QApplication.instance()
        screen = app.screenAt(QCursor.pos()) or app.primaryScreen()
        avail = screen.availableGeometry()

        preferred_w, preferred_h = 2200, 1320
        margin = 40

        if preferred_w > avail.width() or preferred_h > avail.height():
            self.move(avail.topLeft())
            self.showMaximized()
            return

        # Иначе — подгоняем размер под экран и центрируем
        w = min(preferred_w, avail.width() - margin)
        h = min(preferred_h, avail.height() - margin)
        self.resize(w, h)

        rect = self.frameGeometry()
        rect.moveCenter(avail.center())
        self.move(rect.topLeft())

    def closeEvent(self, event):
        self._save_current_queue_state()
        self._sync_queues_with_ui() 
        self.memory_panel.save_instructions_to_disk()

        self.queue_manager.save_current_state()

        if hasattr(self, 'controls_widget') and hasattr(self.controls_widget, "split_results_sw"):
            self.app_settings["global_split_results"] = self.controls_widget.split_results_sw.isChecked()
        self._save_settings()

        if hasattr(self, 'tracker'):
            self.tracker.stop()
        
        try:
            self.controller.cleanup()
        finally:
            event.accept()